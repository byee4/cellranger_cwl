#!/usr/bin/env python

from openpyxl import load_workbook
from argparse import ArgumentParser
import sys
import re
import pandas as pd
from StringIO import StringIO

IGM_SAMPLE_METADATA_MAPPER = {
    'quantification':'Quantification Method',
    'library_prep':'Library Prep Method',
    'uid':'Sample Code',
    'size':'Library Size (bp)',
    'volume':'Volume (ml)',
    'concentration':'Conc. (nM)',
    'total_ng':'Total ng',
    'comment':'Comment',
    'index_1':'Index 1 (Name)',
    'index_2':'Index 2 (Name)',
    'sample_name':'Sample Name'
}


def get_max_lines(sheet):
    """
    returns the max dimension (either row or col) of a sheet.
    """
    dim = sheet.calculate_dimension()
    return max(
        [int(i) for i in re.findall('[\d]+', dim)]
    )


def fill_sample_details_from_string(st, sheet, igm_sample_metadata_mapper,
                                    sep='\t'):
    # stio = StringIO(st)
    df = pd.read_csv(StringIO(st), sep=sep)
    df.columns = [igm_sample_metadata_mapper[c] for c in df.columns] # transform
    offset = 1
    fields = df.columns
    print('fields:', fields)
    for _, sample_id in df.iterrows():
        offset += 1
        for field in fields:
            row, col = get_tofill(sheet, field, offset=offset,
                                  horizontal=False)
            sheet.cell(row=row, column=col).value = sample_id[field]
            print('adding {} to {}, {}'.format(sample_id[field], row, col))
    return sheet

def get_tofill(sheet, label, row_start=1, col_start=1, horizontal=True, offset=1):
    """
    returns the cell that needs to be filled given a field label.
    params:

    sheet: openpyxl.Worksheet
    label: Worksheet cell value
    row_start: int
        topmost value in workbook (1 = 1)
    col_start: int
        leftmost value in workbook (A = 1)
    horizontal:
        if True, return position of label + offset (to the right, horizontally)
        if False, return position of label + offset (downward, vertically)
    offset: int
    """
    for i in range(row_start, get_max_lines(sheet) + 1):
        for j in range(col_start, get_max_lines(sheet) + 1):
            field = sheet.cell(row=i, column=j).value
            if label == field:
                if horizontal:
                    return i, (j + offset)
                else:
                    return (i + offset), j
    print("found nothing for label: {}".format(label))
    return -1, -1


def string2igm(st, xlsx_template, output_file, active_sheet):

    wb = load_workbook(
        filename=xlsx_template
    )
    sheet = wb[active_sheet]
    fill_sample_details_from_string(st, sheet, IGM_SAMPLE_METADATA_MAPPER)
    wb.save(output_file)


def main():
    # Setup argument parser
    parser = ArgumentParser()

    parser.add_argument(
        "--inputstring",
        dest="inputstring",
        help="input arguments as string",
        required=True
    )
    parser.add_argument(
        "--xlsx",
        dest="xlsx",
        help="template xlsx file",
        required=True
    )
    parser.add_argument(
        "--output",
        dest="output",
        help="output xlsx file",
        required=True
    )
    parser.add_argument(
        "--sheet",
        dest="sheet",
        help="the active sheet containing the information",
        default="Sample Information",
        required=False
    )
    try:
        args = parser.parse_args()
    except:
        parser.print_help()
        sys.exit(0)
    print("inputstring: ",args.inputstring)
    inputstring = args.inputstring
    xlsx = args.xlsx
    output = args.output
    active_sheet = args.sheet
    string2igm(inputstring, xlsx, output, active_sheet)


if __name__ == "__main__":
    main()