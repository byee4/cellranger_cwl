#!/usr/bin/env python

from openpyxl import load_workbook
from StringIO import StringIO
from argparse import ArgumentParser
import sys
import re
import pandas as pd

def get_tofill(sheet, label, row_start=1, col_start=1, horizontal=True, offset=1):
    """
    returns the cell that needs to be filled given a field label.
    params:

    sheet: openpyxl.Worksheet
    label: Worksheet cell value
    row_start: int
        topmost value in workbook (1 = 1)
    col_start: int
        leftmost value in workbook (A = 1)
    horizontal:
        if True, return position of label + offset (to the right, horizontally)
        if False, return position of label + offset (downward, vertically)
    offset: int
    """
    for i in range(row_start, get_max_lines(sheet) + 1):
        for j in range(col_start, get_max_lines(sheet) + 1):
            field = sheet.cell(row=i, column=j).value
            if label == field:
                if horizontal:
                    return i, (j + offset)
                else:
                    return (i + offset), j
    print("found nothing for label: {}".format(label))
    return -1, -1

def get_max_lines(sheet):
    """
    returns the max dimension (either row or col) of a sheet.
    """
    dim = sheet.calculate_dimension()
    return max(
        [int(i) for i in re.findall('[\d]+',dim)]
    )

def fill_expt_details_from_field(
    date, institute, pi_name, pi_email, member_of_mcc,
    member_of_drc, contact_name, contact_email, contact_phone_number,
    index_number_or_po, sequencing_platform, type_of_run,
    number_of_cycles, run_bioanalyzer_tape_station, perform_qpcr,
    phix_spike_in_requested, percentage_of_phix_if_applicable,
    indexing_system, total_number_of_lanes, total_number_of_samples,
    sheet
):
    fields = {
        'Date of Sample Submission':date,
        'Institute/Company Name': institute,
        'PI Name': pi_name,
        'PI Email': pi_email,
        'Member of MCC':member_of_mcc,
        'Member of DRC':member_of_drc,
        'Contact Name':contact_name,
        'Contact Email':contact_email,
        'Contact Phone Number':contact_phone_number,
        'Index number or PO':index_number_or_po,
        'Sequencing Platform':sequencing_platform,
        'Type of Run':type_of_run,
        'Number of cycles':number_of_cycles,
        'Run Bioanalyzer/Tape Station':run_bioanalyzer_tape_station,
        'Perform qPCR':perform_qpcr,
        'PhiX Spike In Requested?':phix_spike_in_requested,
        'Percentage of PhiX if applicable':percentage_of_phix_if_applicable,
        'Indexing System':indexing_system,
        'Total number of lanes':total_number_of_lanes,
        'Total number of samples':total_number_of_samples
    }
    for key, value in fields.iteritems():
        row, col = get_tofill(
            sheet, key, offset=1, horizontal=True
        )
        if row == -1 and col == -1:
            print(key, value)
        sheet.cell(row=row, column=col).value = value

def toigm(
    date, institute, pi_name, pi_email, member_of_mcc,
    member_of_drc, contact_name, contact_email, contact_phone_number,
    index_number_or_po, sequencing_platform, type_of_run,
    number_of_cycles, run_bioanalyzer_tape_station, perform_qpcr,
    phix_spike_in_requested, percentage_of_phix_if_applicable,
    indexing_system, total_number_of_lanes, total_number_of_samples,
    xlsx_template, output_file, active_sheet
):
    wb = load_workbook(
        filename=xlsx_template
    )
    sheet = wb[active_sheet]
    fill_expt_details_from_field(
        date, institute, pi_name, pi_email, member_of_mcc,
        member_of_drc, contact_name, contact_email, contact_phone_number,
        index_number_or_po, sequencing_platform, type_of_run,
        number_of_cycles, run_bioanalyzer_tape_station, perform_qpcr,
        phix_spike_in_requested, percentage_of_phix_if_applicable,
        indexing_system, total_number_of_lanes, total_number_of_samples,
        sheet
    )
    wb.save(output_file)


def main():
    # Setup argument parser
    parser = ArgumentParser()
    parser.add_argument(
        "--date", dest='date'
    )
    parser.add_argument(
        "--institute", dest='institute'
    )
    parser.add_argument(
        "--pi_name", dest='pi_name'
    )
    parser.add_argument(
        "--pi_email", dest='pi_email'
    )
    parser.add_argument(
        "--member_of_mcc", dest='member_of_mcc'
    )
    parser.add_argument(
        "--member_of_drc", dest='member_of_drc'
    )
    parser.add_argument(
        "--contact_name", dest='contact_name'
    )
    parser.add_argument(
        "--contact_email", dest='contact_email'
    )
    parser.add_argument(
        "--contact_phone_number", dest='contact_phone_number'
    )
    parser.add_argument(
        "--index_number_or_po", dest='index_number_or_po'
    )
    parser.add_argument(
        "--sequencing_platform", dest='sequencing_platform'
    )
    parser.add_argument(
        "--type_of_run", dest='type_of_run'
    )
    parser.add_argument(
        "--number_of_cycles", dest='number_of_cycles'
    )
    parser.add_argument(
        "--run_bioanalyzer_tape_station", dest='run_bioanalyzer_tape_station'
    )
    parser.add_argument(
        "--perform_qpcr", dest='perform_qpcr'
    )
    parser.add_argument(
        "--phix_spike_in_requested", dest='phix_spike_in_requested'
    )
    parser.add_argument(
        "--percentage_of_phix_if_applicable", dest='percentage_of_phix_if_applicable'
    )
    parser.add_argument(
        "--indexing_system", dest='indexing_system'
    )
    parser.add_argument(
        "--total_number_of_lanes", dest='total_number_of_lanes'
    )
    parser.add_argument(
        "--total_number_of_samples", dest='total_number_of_samples'
    )
    parser.add_argument(
        "--xlsx",
        dest="xlsx",
        help="template xlsx file",
        required=True
    )
    parser.add_argument(
        "--output",
        dest="output",
        help="output xlsx file",
        required=True
    )
    parser.add_argument(
        "--sheet",
        dest="sheet",
        help="the active sheet containing the information",
        default="Sample Information",
        required=False
    )


    try:
        args = parser.parse_args()
    except:
        parser.print_help()
        sys.exit(0)

    toigm(
        args.date, args.institute, args.pi_name, args.pi_email, args.member_of_mcc,
        args.member_of_drc, args.contact_name, args.contact_email, args.contact_phone_number,
        args.index_number_or_po, args.sequencing_platform, args.type_of_run,
        args.number_of_cycles, args.run_bioanalyzer_tape_station, args.perform_qpcr,
        args.phix_spike_in_requested, args.percentage_of_phix_if_applicable,
        args.indexing_system, args.total_number_of_lanes, args.total_number_of_samples,
        args.xlsx, args.output, args.sheet
    )


if __name__ == "__main__":
    main()